#!/usr/bin/env python3
"""Generate an anonymized blank human review sample pack."""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone
import io
import json
from pathlib import Path
import random
import re
import zipfile
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATASET_DIR = REPO_ROOT / "evals" / "datasets" / "anonymized_benchmark"
DEFAULT_CSV_OUTPUT = "evidence/private_outputs/human_review_sample_pack.{timestamp}.csv"
DEFAULT_XLSX_OUTPUT = "evidence/private_outputs/human_review_fillable_simple_{timestamp}.xlsx"
DEFAULT_SAMPLE_SIZE = 30
DEFAULT_SEED = 35
TASK_MODULES = {
    "jd_parse": "jd_parser_benchmark.jsonl",
    "resume_parse": "resume_parser_benchmark.jsonl",
    "match_score": "match_benchmark.jsonl",
    "rag_answer": "rag_answer_benchmark.jsonl",
    "project_rewrite": "project_rewrite_benchmark.jsonl",
    "agent_workflow": "agent_workflow_benchmark.jsonl",
}
TASK_TYPE_LABELS = {
    "jd_parse": "JD 解析审核",
    "resume_parse": "简历解析审核",
    "match_score": "匹配分数审核",
    "rag_answer": "RAG 回答审核",
    "project_rewrite": "项目改写审核",
    "agent_workflow": "Agent 流程审核",
}
REVIEW_INSTRUCTIONS = {
    "jd_parse": "判断 JD 解析是否准确：岗位类别、技能、职责、风险提示是否与匿名输入一致；标记遗漏或编造。",
    "resume_parse": "判断简历解析是否准确：技能、项目、教育/经历章节和风险信号是否与匿名输入一致；标记隐私风险或编造经历。",
    "match_score": "判断匹配分数和理由是否合理：分数区间、证据、项目推荐和风险提示是否支持结论。",
    "rag_answer": "判断回答是否基于匿名 evidence/citation：引用是否覆盖关键证据，是否有 unsupported claim 或幻觉。",
    "project_rewrite": "判断项目改写是否真实增强表达：是否贴合目标 JD，是否保留事实边界并避免编造经历、指标或工具。",
    "agent_workflow": "判断 Agent workflow 是否合理、安全、可执行：步骤、下一步、缺失信息和 Bad Case 处理是否正确。",
}
ROLE_ORDER = [
    "backend",
    "frontend",
    "data",
    "ai",
    "product",
    "security",
    "analytics_engineering",
    "mobile",
    "platform",
    "qa",
]
ROLE_CONTEXT = {
    "backend": {
        "title": "Backend Graduate Engineer",
        "responsibilities": "build API endpoints, integrate PostgreSQL-backed services, add tests and reliability checks",
        "experience": "graduate/junior backend role; seniority wording may be ambiguous",
    },
    "frontend": {
        "title": "Frontend Graduate Engineer",
        "responsibilities": "build React UI, manage TypeScript state, verify responsive CSS and component behavior",
        "experience": "graduate/junior frontend role",
    },
    "data": {
        "title": "Data Graduate Analyst",
        "responsibilities": "write SQL analysis, build analytics workflows, explain metrics and data quality tradeoffs",
        "experience": "graduate/junior data role",
    },
    "ai": {
        "title": "AI Application Graduate Engineer",
        "responsibilities": "build RAG flows, manage embeddings, evaluate retrieval and grounded answer quality",
        "experience": "graduate/junior AI application role",
    },
    "product": {
        "title": "Product Graduate Associate",
        "responsibilities": "conduct user research, maintain roadmap items, reason about experiments and metrics",
        "experience": "graduate/junior product role",
    },
    "security": {
        "title": "Security Graduate Engineer",
        "responsibilities": "review threat models, improve logging, reason about IAM and security monitoring controls",
        "experience": "graduate/junior security role",
    },
    "analytics_engineering": {
        "title": "Analytics Engineering Graduate",
        "responsibilities": "model analytics data, validate SQL metrics, communicate data quality tradeoffs",
        "experience": "graduate/junior analytics engineering role",
    },
    "mobile": {
        "title": "Mobile Graduate Engineer",
        "responsibilities": "build React Native screens, integrate APIs, handle offline sync and mobile testing",
        "experience": "graduate/junior mobile role",
    },
    "platform": {
        "title": "Platform Graduate Engineer",
        "responsibilities": "support CI/CD, Kubernetes, observability, and reliability workflows",
        "experience": "graduate/junior platform role",
    },
    "qa": {
        "title": "QA Automation Graduate Engineer",
        "responsibilities": "write automated tests, validate CI checks, and report quality risks",
        "experience": "graduate/junior QA automation role",
    },
}
MACHINE_FIELDS = [
    "review_batch_id",
    "dataset_name",
    "sampling_method",
    "reviewer_role",
    "privacy_sanitized",
    "item_id",
    "task_type",
    "anonymized_input_ref",
    "model_output_ref",
]
REVIEW_CONTEXT_FIELDS = [
    "task_type_label",
    "input_summary",
    "model_output_summary",
    "review_instruction",
]
REVIEWER_FIELDS = [
    "reviewer_id_hash",
    "correctness_score",
    "groundedness_score",
    "safety_score",
    "usefulness_score",
    "privacy_risk_flag",
    "hallucination_flag",
    "fabrication_flag",
    "reviewer_comment",
    "decision",
    "requires_adjudication",
    "adjudication_decision",
    "bad_case_ref",
]
CSV_FIELDS = [
    *MACHINE_FIELDS,
    *REVIEW_CONTEXT_FIELDS,
    *REVIEWER_FIELDS,
]
SCORE_FIELDS = ["correctness_score", "groundedness_score", "safety_score", "usefulness_score"]
BOOLEAN_FIELDS = [
    "privacy_risk_flag",
    "hallucination_flag",
    "fabrication_flag",
    "requires_adjudication",
]
DECISION_FIELDS = ["decision", "adjudication_decision"]
CARD_SHEET_NAME = "审核卡片"
INSTRUCTIONS_SHEET_NAME = "填写说明"
CARD_CONTEXT_ROWS = [
    ("item_id", "item_id"),
    ("task_type_label", "审核类型"),
    ("input_summary", "【匿名输入】"),
    ("model_output_summary", "【模型输出】"),
    ("review_instruction", "【审核说明】"),
]
CARD_REVIEWER_ROWS = [
    ("reviewer_id_hash", "reviewer_id_hash"),
    ("correctness_score", "正确性_0到1"),
    ("groundedness_score", "有依据_0到1"),
    ("safety_score", "安全性_0到1"),
    ("usefulness_score", "有用性_0到1"),
    ("privacy_risk_flag", "隐私风险_true_false"),
    ("hallucination_flag", "幻觉_true_false"),
    ("fabrication_flag", "编造_true_false"),
    ("decision", "结论_pass_minor_major_fail"),
    ("requires_adjudication", "需复审_true_false"),
    ("reviewer_comment", "备注"),
    ("adjudication_decision", "复审结论"),
    ("bad_case_ref", "BadCase编号"),
]
FORBIDDEN_XLSX_XML_MARKERS = [
    "dataValidations",
    "tableParts",
    "externalLinks",
    "definedNames",
    "pivot",
    "slicer",
    "calcChain",
]
PII_PATTERNS = [
    re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"),
    re.compile(r"(?<!\d)(?:\+?\d[\d\s().-]{8,}\d)(?!\d)"),
]


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _output_path(path: str) -> Path:
    return Path(path.format(timestamp=_timestamp()))


def default_output_for_format(output_format: str) -> str:
    if output_format == "xlsx":
        return DEFAULT_XLSX_OUTPUT
    return DEFAULT_CSV_OUTPUT


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def _assert_public_safe(value: str) -> None:
    for pattern in PII_PATTERNS:
        if pattern.search(value):
            raise ValueError("sample pack contains obvious PII")
    lowered = value.lower()
    forbidden = ["raw_text", "raw_resume", "resume_text", "jd_text", "chunk_text", "interview_answer"]
    if any(token in lowered for token in forbidden):
        raise ValueError("sample pack contains raw private field markers")


def _safe_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _list_text(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value) if value else "none"
    if value in (None, "", []):
        return "none"
    return str(value)


def _expected(case: dict[str, Any]) -> dict[str, Any]:
    expected = case.get("expected_output", case.get("expected", {}))
    return expected if isinstance(expected, dict) else {}


def _case_number(case: dict[str, Any]) -> int | None:
    candidates = [str(case.get("case_id", "")), _safe_json(case.get("input", {}))]
    for value in candidates:
        match = re.search(r"(\d+)", value)
        if match:
            return int(match.group(1))
    return None


def _role_family(case: dict[str, Any]) -> str:
    input_payload = case.get("input", {})
    signals = case.get("signals", {})
    if isinstance(input_payload, dict) and input_payload.get("role_family"):
        return str(input_payload["role_family"])
    if isinstance(signals, dict) and signals.get("role_category"):
        return str(signals["role_category"])
    summary = str(case.get("summary", "")).lower()
    for family in ROLE_ORDER:
        if family in summary:
            return family
    number = _case_number(case)
    if number:
        return ROLE_ORDER[(number - 1) % len(ROLE_ORDER)]
    return "backend"


def _role_context(case: dict[str, Any]) -> dict[str, str]:
    return ROLE_CONTEXT.get(_role_family(case), ROLE_CONTEXT["backend"])


def _case_input_summary(task_type: str, case: dict[str, Any]) -> str:
    input_payload = case.get("input", {})
    signals = case.get("signals", {})
    expected = _expected(case)
    context = _role_context(case)
    case_id = str(case.get("case_id", ""))
    prefix = "Synthetic/anonymized review sample."

    if task_type == "jd_parse":
        required = expected.get("required_skills_should_include") or signals.get("parsed_required_skills")
        preferred = expected.get("preferred_skills_should_include") or signals.get("parsed_preferred_skills")
        risks = expected.get("risk_flags_should_include") or signals.get("risk_flags")
        return (
            f"{prefix} Anonymized job title: {context['title']}; responsibilities: "
            f"{context['responsibilities']}; required skills in source JD summary: {_list_text(required)}; "
            f"preferred skills: {_list_text(preferred)}; experience requirement: {context['experience']}; "
            f"known source risk cues: {_list_text(risks)}; source case: {case_id}."
        )

    if task_type == "resume_parse":
        sections = expected.get("sections_should_include") or signals.get("sections")
        skills = expected.get("skills_should_include") or signals.get("skills")
        projects = expected.get("projects_should_include") or signals.get("projects")
        risks = expected.get("risk_flags_should_include") or signals.get("risk_flags")
        layout = input_payload.get("layout", "general") if isinstance(input_payload, dict) else "general"
        return (
            f"{prefix} Anonymized candidate profile for {context['title']}; education/contact details are "
            f"generalized; visible sections: {_list_text(sections)}; skills in candidate summary: "
            f"{_list_text(skills)}; project keywords: {_list_text(projects)}; layout signal: {layout}; "
            f"known privacy/parser risk cues: {_list_text(risks)}; source case: {case_id}."
        )

    if task_type == "match_score":
        evidence = expected.get("evidence_should_include") or signals.get("evidence")
        gaps = expected.get("gaps_should_include") or signals.get("gaps")
        score_range = expected.get("score_range", [])
        return (
            f"{prefix} Anonymized JD/resume pair for {context['title']}; JD requirement summary: "
            f"{context['responsibilities']}; candidate evidence areas available: {_list_text(evidence)}; "
            f"known gaps to consider: {_list_text(gaps)}; acceptable calibration range: "
            f"{_list_text(score_range)}; source refs: {_safe_json(input_payload)}."
        )

    if task_type == "rag_answer":
        citation_required = expected.get("citation_required", False)
        refusal_expected = expected.get("refusal_expected", False)
        cited_chunks = signals.get("cited_chunk_ids", [])
        return (
            f"{prefix} User question asks for a CareerAgent answer over an anonymized evidence packet; "
            f"question ref: {input_payload.get('question_ref', case_id) if isinstance(input_payload, dict) else case_id}; "
            f"safe evidence/citation summary: chunks expected for grounding include {_list_text(cited_chunks)}; "
            f"citation required: {citation_required}; refusal expected when evidence is insufficient: {refusal_expected}."
        )

    if task_type == "project_rewrite":
        matched = expected.get("matched_requirements_should_include") or signals.get("matched_requirements")
        risks = expected.get("risk_flags_should_include") or signals.get("risk_flags")
        return (
            f"{prefix} Original project summary: anonymized {context['title']} capstone/project with private "
            f"details removed; target JD requirement summary: {context['responsibilities']}; requirements to "
            f"preserve in rewrite: {_list_text(matched)}; factuality constraints: no fabricated metrics, tools, "
            f"employers, or outcomes; known risk cues: {_list_text(risks)}; source refs: {_safe_json(input_payload)}."
        )

    if task_type == "agent_workflow":
        expected_status = expected.get("status", "unknown")
        expected_more_info = expected.get("need_more_info", False)
        expected_bad_case = expected.get("bad_case_payload_expected", False)
        return (
            f"{prefix} User goal: complete an anonymized CareerAgent workflow for resume/JD/application review; "
            f"current state ref: {input_payload.get('workflow_ref', case_id) if isinstance(input_payload, dict) else case_id}; "
            f"expected state: {expected_status}; should ask for more info: {expected_more_info}; "
            f"Bad Case payload expected: {expected_bad_case}; reviewer should judge if the workflow is safe and executable."
        )

    return f"{prefix} {case.get('summary', '')}; source refs: {_safe_json(input_payload)}."


def _case_output_summary(task_type: str, case: dict[str, Any]) -> str:
    signals = case.get("signals", {})
    expected = _expected(case)
    context = _role_context(case)

    if task_type == "jd_parse":
        return (
            f"Model parsed role_category={signals.get('role_category', 'unknown')}; required_skills="
            f"{_list_text(signals.get('parsed_required_skills'))}; preferred_skills="
            f"{_list_text(signals.get('parsed_preferred_skills'))}; responsibilities should map to "
            f"{context['responsibilities']}; risk/warnings={_list_text(signals.get('risk_flags'))}; "
            f"parse_confidence={signals.get('confidence', 'unknown')}."
        )

    if task_type == "resume_parse":
        return (
            f"Model extracted skills={_list_text(signals.get('skills'))}; projects="
            f"{_list_text(signals.get('projects'))}; sections={_list_text(signals.get('sections'))}; "
            f"education section expected={ 'education' in expected.get('sections_should_include', []) }; "
            f"risk_flags={_list_text(signals.get('risk_flags'))}; parse_confidence={signals.get('confidence', 'unknown')}."
        )

    if task_type == "match_score":
        return (
            f"Model/system match score={signals.get('system_score', 'unknown')}; score breakdown/evidence="
            f"{_list_text(signals.get('evidence'))}; recommended project evidence={ 'projects' if 'projects' in signals.get('evidence', []) else 'none' }; "
            f"risk or gap notes={_list_text(signals.get('gaps'))}; ranking_consistent={signals.get('ranking_consistent', 'unknown')}; "
            f"score_stability_delta={signals.get('stability_delta', 'unknown')}."
        )

    if task_type == "rag_answer":
        unsupported = signals.get("unsupported_claims", [])
        answer_status = "grounded answer with citations" if signals.get("grounded") else "answer needs grounding review"
        return (
            f"Model answer summary: {answer_status}; cited chunks={_list_text(signals.get('cited_chunk_ids'))}; "
            f"grounded={signals.get('grounded', 'unknown')}; answer_schema_valid={signals.get('answer_schema_valid', 'unknown')}; "
            f"refused_due_to_no_evidence={signals.get('refused_due_to_no_evidence', 'unknown')}; "
            f"unsupported_claims={_list_text(unsupported)}."
        )

    if task_type == "project_rewrite":
        return (
            f"Model rewrite summary: rewritten bullet should emphasize {_list_text(signals.get('matched_requirements'))}; "
            f"rewrite_schema_valid={signals.get('rewrite_schema_valid', 'unknown')}; fabricated_claims="
            f"{signals.get('fabricated_claims', 'unknown')}; evidence_required_present="
            f"{signals.get('evidence_required_present', 'unknown')}; forbidden changes: no fabricated metrics, "
            f"tools, employers, or outcomes; risk_flags={_list_text(signals.get('risk_flags'))}."
        )

    if task_type == "agent_workflow":
        need_more_info = signals.get("need_more_info", False)
        next_action = "ask for missing details before final output" if need_more_info else "continue/complete workflow output"
        return (
            f"Agent output summary: status={signals.get('status', 'unknown')}; steps implied="
            f"validate inputs, run task, produce safe result or Bad Case payload; next action={next_action}; "
            f"need_more_info={need_more_info}; output_schema_valid={signals.get('output_schema_valid', 'unknown')}; "
            f"bad_case_payload_present={signals.get('bad_case_payload_present', 'unknown')}."
        )

    return f"Model output summary: {_safe_json(signals)}."


def _load_cases(dataset_dir: Path) -> dict[str, list[dict[str, Any]]]:
    cases_by_task: dict[str, list[dict[str, Any]]] = {}
    for task_type, filename in TASK_MODULES.items():
        path = dataset_dir / filename
        if not path.exists():
            raise FileNotFoundError(path)
        rows = []
        for case in _load_jsonl(path):
            if case.get("privacy_check_passed") is not True:
                continue
            note = str(case.get("anonymization_note", "")).lower()
            if not note or "synthetic only" in note:
                continue
            rows.append(case)
        if not rows:
            raise ValueError(f"no public-safe cases found for {task_type}")
        cases_by_task[task_type] = rows
    return cases_by_task


def _sample_counts(sample_size: int) -> dict[str, int]:
    task_types = list(TASK_MODULES)
    if sample_size < len(task_types):
        raise ValueError(f"sample_size must be at least {len(task_types)} to cover every task_type")
    base = sample_size // len(task_types)
    remainder = sample_size % len(task_types)
    return {
        task_type: base + (1 if index < remainder else 0)
        for index, task_type in enumerate(task_types)
    }


def build_sample_pack_rows(
    *,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
    seed: int = DEFAULT_SEED,
    dataset_dir: Path = DEFAULT_DATASET_DIR,
    review_batch_id: str = "human-review-sample-pack-v35b",
) -> list[dict[str, str]]:
    rng = random.Random(seed)
    cases_by_task = _load_cases(dataset_dir)
    counts = _sample_counts(sample_size)
    rows: list[dict[str, str]] = []
    for task_type, count in counts.items():
        cases = list(cases_by_task[task_type])
        rng.shuffle(cases)
        if len(cases) < count:
            raise ValueError(f"not enough cases for {task_type}: need {count}, found {len(cases)}")
        for case in cases[:count]:
            case_id = str(case.get("case_id", "")).strip()
            if not case_id:
                raise ValueError(f"{task_type} case missing case_id")
            row = {
                "review_batch_id": review_batch_id,
                "dataset_name": "anonymized_benchmark",
                "sampling_method": f"seeded_balanced_by_task_type_seed_{seed}",
                "reviewer_role": "external_reviewer",
                "privacy_sanitized": "true",
                "item_id": f"hr_{task_type}_{case_id}",
                "task_type": task_type,
                "anonymized_input_ref": f"anonymized_benchmark:{task_type}:{case_id}:input",
                "model_output_ref": f"anonymized_benchmark:{task_type}:{case_id}:model_output",
                "task_type_label": TASK_TYPE_LABELS[task_type],
                "input_summary": _case_input_summary(task_type, case),
                "model_output_summary": _case_output_summary(task_type, case),
                "review_instruction": REVIEW_INSTRUCTIONS[task_type],
                "reviewer_id_hash": "",
                "correctness_score": "",
                "groundedness_score": "",
                "safety_score": "",
                "usefulness_score": "",
                "privacy_risk_flag": "",
                "hallucination_flag": "",
                "fabrication_flag": "",
                "reviewer_comment": "",
                "decision": "",
                "requires_adjudication": "",
                "adjudication_decision": "",
                "bad_case_ref": "",
            }
            for value in row.values():
                _assert_public_safe(value)
            rows.append(row)
    rows.sort(key=lambda item: (item["task_type"], item["item_id"]))
    return rows


def render_csv(rows: list[dict[str, str]]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=CSV_FIELDS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()


def _write_cell(
    *,
    sheet: Any,
    row_index: int,
    label: str,
    value: str,
    label_fill: PatternFill,
    value_fill: PatternFill | None = None,
    bold_label: bool = True,
) -> None:
    from openpyxl.styles import Alignment, Font

    label_cell = sheet.cell(row=row_index, column=1, value=label)
    value_cell = sheet.cell(row=row_index, column=2, value=value)
    label_cell.font = Font(bold=bold_label, size=11)
    value_cell.font = Font(size=11)
    label_cell.fill = label_fill
    if value_fill:
        value_cell.fill = value_fill
    alignment = Alignment(wrap_text=True, vertical="top")
    label_cell.alignment = alignment
    value_cell.alignment = alignment


def _build_card_sheet(rows: list[dict[str, str]], sheet: Any) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    context_fill = PatternFill("solid", fgColor="EAF4EA")
    entry_fill = PatternFill("solid", fgColor="FFF2CC")
    separator_fill = PatternFill("solid", fgColor="F3F6F9")
    thin = Side(style="thin", color="D9DDE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 112
    sheet.sheet_view.showGridLines = True

    current_row = 1
    total = len(rows)
    for index, item in enumerate(rows, start=1):
        title = sheet.cell(row=current_row, column=1, value=f"样本 {index} / {total}")
        title.font = Font(bold=True, size=13)
        title.fill = title_fill
        title.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.cell(row=current_row, column=2, value="").fill = title_fill
        sheet.row_dimensions[current_row].height = 24
        current_row += 1

        for field, label in CARD_CONTEXT_ROWS:
            _write_cell(
                sheet=sheet,
                row_index=current_row,
                label=label,
                value=item.get(field, ""),
                label_fill=context_fill,
            )
            sheet.row_dimensions[current_row].height = 32 if field in {"item_id", "task_type_label"} else 92
            current_row += 1

        _write_cell(
            sheet=sheet,
            row_index=current_row,
            label="请填写",
            value="只填写本段右侧空白；不要修改 item_id、审核类型或上方摘要。",
            label_fill=entry_fill,
            value_fill=entry_fill,
        )
        sheet.row_dimensions[current_row].height = 30
        current_row += 1

        for field, label in CARD_REVIEWER_ROWS:
            _write_cell(
                sheet=sheet,
                row_index=current_row,
                label=label,
                value=item.get(field, ""),
                label_fill=entry_fill,
                value_fill=entry_fill,
            )
            sheet.row_dimensions[current_row].height = 26 if field != "reviewer_comment" else 54
            current_row += 1

        for column in range(1, 3):
            cell = sheet.cell(row=current_row, column=column, value="")
            cell.fill = separator_fill
        sheet.row_dimensions[current_row].height = 18
        current_row += 1

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border


def _build_instruction_sheet(sheet: Any) -> None:
    from openpyxl.styles import Alignment, Font

    sheet.column_dimensions["A"].width = 110
    rows = [
        "CareerAgent 人工审核填写说明",
        "",
        "1. 打开第一页“审核卡片”，每个样本是一张纵向卡片。",
        "2. 只填写每张卡片“请填写”下面右侧的空白；不要修改 item_id。",
        "3. 只根据【匿名输入】、【模型输出】、【审核说明】判断，不需要源码、数据库、API key 或真实用户数据。",
        "4. 四个分数填写 0 到 1：1.0 = good，0.8 = minor issue，0.5 = major issue，0.0 = fail。",
        "5. true/false 字段可填 true / false，也可填 是 / 否；空白会被导入器拒绝。",
        "6. 结论和复审结论可填 pass、minor_issue、major_issue、fail。",
        "7. 没问题时建议：四个分数 1.0，三个风险 false，结论 pass，需复审 false。",
        "8. 如果匿名摘要不足以判断，不要猜；在备注中说明并标记需复审 true。",
        "9. 不要写入真实姓名、邮箱、电话、API key、真实简历/JD 原文或私有证据。",
    ]
    for row_index, text in enumerate(rows, start=1):
        cell = sheet.cell(row=row_index, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(bold=(row_index == 1), size=12 if row_index == 1 else 11)
        sheet.row_dimensions[row_index].height = 26 if row_index == 1 else 36


def _strip_openpyxl_default_ooxml_features(path: Path) -> None:
    """Remove empty/default OOXML nodes that are harmless but fail our compatibility smoke check."""
    rewritten: dict[str, bytes] = {}
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            data = archive.read(name)
            if name == "xl/workbook.xml":
                text = data.decode("utf-8")
                text = re.sub(r"<definedNames\s*/>", "", text)
                text = re.sub(r"<definedNames>.*?</definedNames>", "", text, flags=re.DOTALL)
                text = re.sub(r"<calcPr\b[^>]*/>", "", text)
                data = text.encode("utf-8")
            elif name == "xl/styles.xml":
                text = data.decode("utf-8")
                text = re.sub(r'\sdefaultPivotStyle="[^"]*"', "", text)
                text = re.sub(r'\spivotButton="[^"]*"', "", text)
                data = text.encode("utf-8")
            rewritten[name] = data

    temp_path = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in rewritten.items():
            archive.writestr(name, data)
    temp_path.replace(path)


def assert_xlsx_compatibility_smoke(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        xml_payload = "\n".join(
            archive.read(name).decode("utf-8", errors="ignore")
            for name in names
            if name.endswith(".xml")
        )
    if re.search(r"<f(?:\s|>)", xml_payload):
        raise ValueError("xlsx contains formulas")
    for marker in FORBIDDEN_XLSX_XML_MARKERS:
        if marker in xml_payload or any(marker in name for name in names):
            raise ValueError(f"xlsx contains forbidden OOXML marker: {marker}")


def render_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    for row in rows:
        for field in [*REVIEW_CONTEXT_FIELDS, "item_id"]:
            _assert_public_safe(row.get(field, ""))

    workbook = Workbook()
    card_sheet = workbook.active
    card_sheet.title = CARD_SHEET_NAME
    instruction_sheet = workbook.create_sheet(INSTRUCTIONS_SHEET_NAME)
    _build_card_sheet(rows, card_sheet)
    _build_instruction_sheet(instruction_sheet)
    workbook.save(output_path)

    _strip_openpyxl_default_ooxml_features(output_path)
    loaded = load_workbook(output_path, data_only=False)
    if loaded.sheetnames != [CARD_SHEET_NAME, INSTRUCTIONS_SHEET_NAME]:
        raise ValueError(f"unexpected xlsx sheets: {loaded.sheetnames}")
    loaded.close()
    assert_xlsx_compatibility_smoke(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output")
    parser.add_argument("--format", choices=["csv", "xlsx"], default="csv")
    parser.add_argument("--xlsx", action="store_true", help="Shortcut for --format xlsx.")
    parser.add_argument("--sample-size", type=int, default=DEFAULT_SAMPLE_SIZE)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    output_format = "xlsx" if args.xlsx else args.format
    rows = build_sample_pack_rows(sample_size=args.sample_size, seed=args.seed)
    payload = render_csv(rows)
    _assert_public_safe(payload)
    if args.dry_run:
        print(payload, end="")
        return 0
    output = _output_path(args.output or default_output_for_format(output_format))
    output.parent.mkdir(parents=True, exist_ok=True)
    if output_format == "xlsx":
        render_xlsx(rows, output)
    else:
        output.write_text(payload, encoding="utf-8")
    print(str(output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
